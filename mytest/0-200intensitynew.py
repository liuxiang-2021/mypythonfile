# -*- coding: utf-8 -*-
# @Time    : 2021/3/4 20:45
# @Author  : liuxiang
# @FileName: 0-200intensitynew.py
# @Software: PyCharm
# @email    ：liuxiang@hesaitech.com

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import random
import openpyxl
from openpyxl import Workbook
import time

nDistances = 20         # 分析距离的数目 （CSV文件个数）
distanceMin = 10        # 最小距离
distanceMax = 200
stepDistance = 10

# 定义画散点图的函数
def draw_scatter(s):
    """
    #:param n: 点的数量，整数
    :param s:点的大小，整数
    :return: None
    """
    # 创建画图窗口
    fig = plt.figure()
    # 将画图窗口分成1行1列，选择第一块区域作子图
    ax1 = fig.add_subplot(1, 1, 1)
    # 设置标题
    ax1.set_title('Intensity Analysis')
    # 设置横坐标名称
    ax1.set_xlabel('Distance(m)')
    # 设置纵坐标名称
    ax1.set_ylabel('Intensity(%)')

    xlist = []
    ylist = []
    tmplist = []
    # 加载数据
    df = pd.read_csv("intensity.csv")
    # 获取列的命长
    df_title = df.columns.values.tolist()
    # 处理每列数据
    for i in range(nDistances):
        tmplist = list(df[df_title[i]])
        nLines = len(tmplist)
        ylist.append(tmplist)
        xlist.append([stepDistance*i+distanceMin]*nLines)
        for j in range(nLines):
            xlist[i][j] += random.normalvariate(stepDistance/2,stepDistance/3)
            ylist[i][j] += random.normalvariate(0.5,0.5)+0.5
        # 画散点图
        ax1.scatter(xlist[i], ylist[i], s=s, c='blue',marker='.')



    # 画直线图
    #ax1.plot(xlist[0], ylist[0], c='b', ls='--')
    # 调整横坐标的上下界
    plt.xlim(xmax=200, xmin=0)
    plt.ylim(ymax=30, ymin=0)
    # 显示
    plt.show()

# 定义读文件的函数
def readcsvs():
    # 创建xlsx
    wb = Workbook()
    wb.save('intensity.xlsx')
    data = openpyxl.load_workbook('intensity.xlsx')
    #创建sheet名称为‘intensity’
    ws = data.create_sheet(0)
    ws.title='intensity'
    table = data.get_sheet_by_name('intensity')

    # 文件名为10.csv   15.csv   20.csv  ...
    a = distanceMin
    while(a<=distanceMax):
        df = pd.read_csv(str(a)+'.csv')
        df1 = df[(df["intensity"] >= 0) & (df["laser_id"] != 47)]
        reflectanceList = list(df1['intensity'])
        table.cell(1, (a - distanceMin+stepDistance) / stepDistance).value = a
        for nRow in range(len(reflectanceList)):
            table.cell(nRow+2,(a - distanceMin+stepDistance)/stepDistance).value = reflectanceList[nRow]
        a += stepDistance
    data.save('intensity.xlsx')
    #xls转化为csv
    data_xls = pd.read_excel('intensity.xlsx', sheet_name='intensity',index_col=0)
    data_xls.to_csv('intensity.csv', encoding='utf-8')


# 主模块
if __name__ == "__main__":
    # 运行
    readcsvs()
    time.sleep(2)
    draw_scatter(s=1)
